from paddleocr import PaddleOCR
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

os.environ['FLAGS_selected_devices'] = 'cpu'

ocr = PaddleOCR(lang='ch', use_gpu=False, use_angle_cls=True)

# Process all three images
images = [
    ('D:\\2024年度工作日志和备忘录\\数字化转型产品\\4.0 同事组\\5.0 邓\\2026\\0430\\微信图片_20260424134130_153_109.png', '中国供应链百强'),
    ('D:\\2024年度工作日志和备忘录\\数字化转型产品\\4.0 同事组\\5.0 邓\\2026\\0430\\微信图片_20260424134130_154_109.png', '全球零部件供应商百强'),
    ('D:\\2024年度工作日志和备忘录\\数字化转型产品\\4.0 同事组\\5.0 邓\\2026\\0430\\微信图片_20260424134419_156_109.png', '动力电池装机量'),
]

def extract_texts(img_path):
    result = ocr.ocr(img_path)
    texts = []
    for line in result[0]:
        text = line[1][0]
        score = line[1][1]
        texts.append((text, score))
    return texts

# Process image 3 (battery data) - already analyzed above
battery_texts = [
    ('2025年6月国内动力电池企业装车量前十五名', 0.99),
    ('企业名称', 1.00),
    ('与上月比占比变', 0.99),
    ('装车量（GWh）', 1.00),
    ('占比', 1.00),
    ('序号', 1.00),
    ('化/百分点', 0.98),
    ('宁德时代', 0.99),
    ('25.41', 1.00),
    ('43.67%', 1.00),
    ('0.80', 1.00),
    ('比亚迪', 1.00),
    ('12.49', 1.00),
    ('21.47%', 1.00),
    ('-1.02', 1.00),
    ('中创新航', 1.00),
    ('4.39', 1.00),
    ('7.55%', 1.00),
    ('0.02', 1.00),
    ('国轩高科', 1.00),
    ('2.92', 1.00),
    ('5.02%', 1.00),
    ('-0.24', 0.99),
    ('亿纬锂能', 0.99),
    ('2.53', 1.00),
    ('4.35%', 1.00),
    ('0.69', 1.00),
    ('欣旺达', 1.00),
    ('2.17', 1.00),
    ('3.72%', 0.99),
    ('-0.20', 0.99),
    ('正力新能', 1.00),
    ('1.28', 1.00),
    ('2.21%', 1.00),
    ('-0.98', 0.99),
    ('瑞浦兰钧', 1.00),
    ('1.28', 1.00),
    ('2.20%', 1.00),
    ('0.04', 1.00),
    ('蜂巢能源', 1.00),
    ('1.20', 1.00),
    ('2.06%', 0.90),
    ('0.02', 1.00),
    ('因湃电池', 1.00),
    ('1.08', 1.00),
    ('1.86%', 1.00),
    ('1.26', 1.00),
    ('LG新能源', 1.00),
    ('0.94', 1.00),
    ('1.61%', 1.00),
    ('0.42', 1.00),
    ('极电新能源', 1.00),
    ('0.68', 1.00),
    ('1.17%', 1.00),
    ('-0.38', 0.99),
    ('楚能新能源', 1.00),
    ('0.44', 1.00),
    ('0.75%', 1.00),
    ('0.01', 1.00),
    ('多氟多', 0.99),
    ('0.25', 1.00),
    ('0.43%', 1.00),
    ('-0.34', 0.99),
    ('远航锦锂', 1.00),
    ('0.24', 1.00),
    ('0.42%', 1.00),
    ('-0.01', 0.99),
]

# Parse battery data
battery_data = []
skip_next = False
for i, (text, score) in enumerate(battery_texts):
    if skip_next:
        skip_next = False
        continue
    if text.isdigit() and 1 <= int(text) <= 15:
        num = int(text)
        # Next items should be company, capacity, percentage, change
        if i + 4 < len(battery_texts):
            company = battery_texts[i + 1][0]
            capacity = battery_texts[i + 2][0]
            percentage = battery_texts[i + 3][0]
            change = battery_texts[i + 4][0]
            battery_data.append({
                '序号': num,
                '公司简称': company,
                '装车量(GWh)': capacity,
                '占比': percentage,
                '同比变化': change
            })
            skip_next = True

print(f"Battery data: {len(battery_data)} entries")
for d in battery_data:
    print(d)

# Process image 1 and 2 with PaddleOCR
print("\nProcessing image 1...")
img1_texts = extract_texts(images[0][0])
print(f"Image 1: {len(img1_texts)} text regions")

print("\nProcessing image 2...")
img2_texts = extract_texts(images[1][0])
print(f"Image 2: {len(img2_texts)} text regions")

# Parse global suppliers data (image 2)
global_data = []
for i, (text, score) in enumerate(img2_texts):
    if text.isdigit() and 1 <= int(text) <= 100:
        num = int(text)
        if i + 4 < len(img2_texts):
            company_cn = img2_texts[i + 1][0]
            company_en = img2_texts[i + 2][0]
            country = img2_texts[i + 3][0]
            revenue = img2_texts[i + 4][0]
            # Filter out garbage
            if company_en and len(company_en) > 2 and not any(c in '，。！？' for c in company_en):
                try:
                    # Try to parse revenue as number
                    rev_num = float(revenue.replace(',', '').replace('(', '').replace(')', '').replace('f', '').replace('e', '').replace('c', '').strip()) if isinstance(revenue, str) else revenue
                    global_data.append({
                        '序号': num,
                        '公司简称': company_cn,
                        '英文名称': company_en,
                        '国家': country,
                        '2024年营收(亿美元)': rev_num
                    })
                except:
                    pass

print(f"\nGlobal data: {len(global_data)} entries")
for d in global_data[:5]:
    print(d)

# Create Excel
wb = openpyxl.Workbook()

# Sheet 1: Global suppliers
ws1 = wb.active
ws1.title = '全球零部件供应商百强'
header1 = ['序号', '公司简称', '英文名称', '国家', '2024年营收(亿美元)']
for col, h in enumerate(header1, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row_idx, row_data in enumerate(global_data, 2):
    for col_idx, key in enumerate(['序号', '公司简称', '英文名称', '国家', '2024年营收(亿美元)'], 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 40
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 22

# Sheet 2: Battery data
ws2 = wb.create_sheet('动力电池装机量')
header2 = ['序号', '公司简称', '装车量(GWh)', '占比', '同比变化']
for col, h in enumerate(header2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row_idx, row_data in enumerate(battery_data, 2):
    for col_idx, key in enumerate(['序号', '公司简称', '装车量(GWh)', '占比', '同比变化'], 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 12

# Sheet 3: Supply chain (image 1 - partially parseable)
ws3 = wb.create_sheet('中国供应链百强')
header3 = ['序号', '公司简称', '2024年营收(亿元)', '地区']
for col, h in enumerate(header3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Extract from image 1 - look for patterns
china_data = []
for i, (text, score) in enumerate(img1_texts):
    if text.isdigit() and 1 <= int(text) <= 100:
        num = int(text)
        if i + 3 < len(img1_texts):
            company = img1_texts[i + 1][0]
            # Skip if it looks like a number or short code
            if company and len(company) > 3 and not company.isdigit():
                revenue = img1_texts[i + 2][0] if i + 2 < len(img1_texts) else ''
                region = img1_texts[i + 3][0] if i + 3 < len(img1_texts) else ''
                # Try to clean revenue
                try:
                    if '亿' in str(revenue) or any(c.isdigit() for c in str(revenue)):
                        china_data.append({
                            '序号': num,
                            '公司简称': company,
                            '2024年营收(亿元)': revenue,
                            '地区': region
                        })
                except:
                    pass

for row_idx, row_data in enumerate(china_data, 2):
    for col_idx, key in enumerate(['序号', '公司简称', '2024年营收(亿元)', '地区'], 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 12

print(f"\nChina supply chain: {len(china_data)} entries")

# Save
output_path = r'D:\2024年度工作日志和备忘录\数字化转型产品\4.0 同事组\5.0 邓\2026\0430\汽车产业公司简称对照表_PaddleOCR.xlsx'
wb.save(output_path)
print(f'\nExcel saved to: {output_path}')
